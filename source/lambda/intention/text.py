import boto3
from openpyxl import load_workbook
from io import BytesIO
import logging

# 配置日志记录器
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def excel():
    # 初始化 S3 客户端
    s3_client = boto3.client('s3')

    # 定义 S3 存储桶名称和文件的路径（Key）
    bucket_name = 'ai-customer-service-apiconstructllmbotdocumentsfc4-kq7pcgbahxwk'
    file_key = 'intentions/Admin/qalist_0822.xlsx'

    try:
        # 从 S3 获取文件内容
        response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        file_content = response['Body'].read()
        print(file_content[:1000])  
        logger.info(f"File content length: {len(file_content)}")

        # 使用 BytesIO 将文件内容作为字节流加载
        excel_file = BytesIO(file_content)
        print(f'=======type is {type(excel_file)}')
        print(f'=======excel_file is {excel_file.getvalue()[:1000]}')  # 

        # 使用 openpyxl 加载 Excel 文件
        workbook = load_workbook(filename="./qalist_0822.xlsx")

        # 检查工作簿中的工作表
        sheet_names = workbook.sheetnames
        logger.info(f"Sheet names: {sheet_names}")

        # 获取第一个工作表并读取内容
        if sheet_names:
            sheet = workbook[sheet_names[0]]
            logger.info(f"First sheet: {sheet.title}")
        
            # 打印工作表中的所有数据
            for row in sheet.iter_rows(values_only=True):
                logger.info(row)
        else:
            logger.info("No sheets found in the workbook.")
    except Exception as e:
        logger.error(f"An error occurred: {e}")

if __name__ == "__main__":
    logger.info("Starting the script")
    excel()

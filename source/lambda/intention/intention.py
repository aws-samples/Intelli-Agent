
import hashlib
import json
import os
import re
import time
from typing import List
import boto3
from openpyxl import load_workbook
from io import BytesIO
from embeddings import get_embedding_info
from botocore.paginate import TokenEncoder
from opensearchpy import NotFoundError, RequestError, helpers, RequestsHttpConnection
import logging
from langchain.embeddings.bedrock import BedrockEmbeddings
from langchain.docstore.document import Document
from langchain_community.vectorstores import OpenSearchVectorSearch
from langchain_community.vectorstores.opensearch_vector_search import (
    OpenSearchVectorSearch,
)
from aos import sm_utils
from requests_aws4auth import AWS4Auth

from aos.aos_utils import LLMBotOpenSearchClient
from constant import (BULK_SIZE,
                      DEFAULT_CONTENT_TYPE,
                      DEFAULT_MAX_ITEMS,
                      DEFAULT_SIZE,
                      DOWNLOAD_RESOURCE,
                      EXECUTION_RESOURCE,
                      INDEX_USED_SCAN_RESOURCE,
                      PRESIGNED_URL_RESOURCE,
                      ModelDimensionMap)

logger = logging.getLogger(__name__)
encoder = TokenEncoder()

s3_bucket_name = os.environ.get("S3_BUCKET")
embedding_model_endpoint = os.environ.get("EMBEDDING_MODEL_ENDPOINT", "")
aosEndpoint = os.environ.get("AOS_ENDPOINT")
kb_enabled = os.environ["KNOWLEDGE_BASE_ENABLED"].lower() == "true"
region = os.environ.get("AWS_REGION", "us-east-1")
bedrock_region = os.environ.get("BEDROCK_REGION", "us-east-1")
aos_domain_name = os.environ.get("AOS_DOMAIN_NAME", "smartsearch")
aos_endpoint = os.environ.get("AOS_ENDPOINT", "")
aos_secret = os.environ.get("AOS_SECRET_NAME", "opensearch-master-user")
intention_table_name = os.getenv("INTENTION_TABLE_NAME", "intention")
chatbot_table_name = os.getenv("CHATBOT_TABLE_NAME", "chatbot")
index_table_name = os.getenv("INDEX_TABLE_NAME", "index")
model_table_name = os.getenv("MODEL_TABLE_NAME", "model")
dynamodb_client = boto3.resource("dynamodb")
intention_table = dynamodb_client.Table(intention_table_name)
index_table = dynamodb_client.Table(index_table_name)
chatbot_table = dynamodb_client.Table(chatbot_table_name)
model_table = dynamodb_client.Table(model_table_name)

sm_client = boto3.client("secretsmanager")
credentials = boto3.Session().get_credentials()
try:
    master_user = sm_client.get_secret_value(
        SecretId=aos_secret)["SecretString"]
    secret = json.loads(master_user)
    username = secret.get("username")
    password = secret.get("password")

    if not aos_endpoint:
        opensearch_client = boto3.client("opensearch")
        response = opensearch_client.describe_domain(
            DomainName=aos_domain_name)
        aos_endpoint = response["DomainStatus"]["Endpoint"]
    aos_client = LLMBotOpenSearchClient(aos_endpoint, (username, password)).client
    awsauth = (username, password)
except sm_client.exceptions.ResourceNotFoundException:
    logger.info("Secret '%s' not found in Secrets Manager", aos_secret)
    aos_client = LLMBotOpenSearchClient(aos_endpoint).client
    awsauth = AWS4Auth(refreshable_credentials=credentials,
                   region=region, service="es")
except sm_client.exceptions.InvalidRequestException:
    logger.info("InvalidRequestException. It might caused by getting secret value from a deleting secret")
    logger.info("Fallback to authentication with IAM")
    aos_client = LLMBotOpenSearchClient(aos_endpoint).client
    awsauth = AWS4Auth(refreshable_credentials=credentials,
                   region=region, service="es")
except Exception as err:
    logger.error("Error retrieving secret '%s': %s", aos_secret, str(err))
    raise

dynamodb_client = boto3.client("dynamodb")
s3_client = boto3.client("s3")
bedrock_client = boto3.client("bedrock-runtime", region_name=bedrock_region)

credentials = boto3.Session().get_credentials()

resp_header = {
    "Content-Type": "application/json",
    "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "*",
}


class OpenSearchIngestionWorker:
    def __init__(
        self,
        docsearch: OpenSearchVectorSearch,
        embedding_model_endpoint: str,
    ):
        self.docsearch = docsearch
        self.embedding_model_endpoint = embedding_model_endpoint

    def aos_ingestion(self, documents: List[Document], index: str) -> None:
        texts = [doc.page_content for doc in documents]
        metadatas = [doc.metadata for doc in documents]
        embeddings_vectors = self.docsearch.embedding_function.embed_documents(
            list(texts)
        )

        if isinstance(embeddings_vectors[0], dict):
            embeddings_vectors_list = []
            metadata_list = []
            for doc_id, metadata in enumerate(metadatas):
                embeddings_vectors_list.append(
                    embeddings_vectors[0]["dense_vecs"][doc_id]
                )
                metadata["embedding_endpoint_name"] = self.embedding_model_endpoint
                metadata_list.append(metadata)
            embeddings_vectors = embeddings_vectors_list
            metadatas = metadata_list
        self.docsearch._OpenSearchVectorSearch__add(
            texts, embeddings_vectors, metadatas=metadatas, index=index
        )


def lambda_handler(event, context):
    logger.info(event)
    authorizer_type = (
        event["requestContext"].get("authorizer", {}).get("authorizerType")
    )
    if authorizer_type == "lambda_authorizer":
        claims = json.loads(event["requestContext"]["authorizer"]["claims"])
        if "use_api_key" in claims:
            group_name = __get_query_parameter(event, "GroupName", "Admin")
        else:
            email = claims["email"]
            # Agree to only be in one group
            group_name = claims["cognito:groups"]
    else:
        logger.error("Invalid authorizer type")
    http_method = event["httpMethod"]
    resource: str = event["resource"]
    if resource == PRESIGNED_URL_RESOURCE:
        input_body = json.loads(event["body"])
        file_name = f"intentions/{group_name}/[{input_body['timestamp']}]{input_body['file_name']}"
        presigned_url = __gen_presigned_url(file_name,
                                            input_body.get(
                                                "content_type", DEFAULT_CONTENT_TYPE),
                                            input_body.get("expiration", 60*60))
        output = {
            "message": "The S3 presigned url is generated",
            "data": {
                "url": presigned_url,
                "s3Bucket": s3_bucket_name,
                "s3Prefix": file_name,
            },

        }
    elif resource.startswith(EXECUTION_RESOURCE):
        if http_method == "POST":
            output = __create_execution(event, context, email, group_name)
        elif http_method == "DELETE":
            output = __delete_execution(event, group_name)
        else:
            if resource == EXECUTION_RESOURCE:
                output = __list_execution(event, group_name)
            else:
                output = __get_execution(event, group_name)
    elif resource == DOWNLOAD_RESOURCE:
        output = __download_template()
    elif resource == INDEX_USED_SCAN_RESOURCE:
        output = __index_used_scan(event, group_name)
    try:
        return {
            "statusCode": 200,
            "headers": resp_header,
            "body": json.dumps(output),
        }
    except Exception as error:
        logger.error("Error: %s", str(error))
        return {
            "statusCode": 500,
            "headers": resp_header,
            "body": json.dumps(f"Error: {str(error)}"),
        }


def __delete_execution(event, group_name):
    input_body = json.loads(event["body"])
    execution_ids = input_body.get("executionIds")
    res = []
    for execution_id in execution_ids:
        index_response = intention_table.get_item(
            Key={
                "groupName": group_name,
                "intentionId": execution_id,
            },
        )
        item = index_response.get('Item')
        if item:
            indexes = item.get("index").split(",")
            details = json.loads(item.get("details"))
            questions = [detail.get("question") for detail in details]
            # delete aos data
            for index in indexes:
                __delete_documents_by_text_set(index, questions)
            # delete intention（ddb）
            intention_table.delete_item(
                Key={
                    "groupName": group_name,
                    "intentionId": execution_id,
                }
            )
    return res

# def __can_be_deleted(execution_id):

#     return False, ""


def __delete_documents_by_text_set(index_name, text_values):
    # Search for the documents based on the "text" field matching any value in text_values set
    search_body = {
        "size": 10000,
        "query": {
            "terms": {
                "text.keyword": list(text_values)  # Convert set to list
            }
        }
    }

    # Perform the search
    try:
        search_result = aos_client.search(
            index=index_name, body=search_body)  # Adjust size if needed
        hits = search_result['hits']['hits']

        # If documents exist, delete them
        if hits:
            for hit in hits:
                doc_id = hit['_id']
                aos_client.delete(index=index_name, id=doc_id)
                logger.info("Deleted document with id %s", doc_id)
    except NotFoundError:
        logger.info("Index is not existed: %s", index_name)


def __get_query_parameter(event, parameter_name, default_value=None):
    if (
        event.get("queryStringParameters")
        and parameter_name in event["queryStringParameters"]
    ):
        return event["queryStringParameters"][parameter_name]
    return default_value


def __gen_presigned_url(object_name: str, content_type: str, expiration: int):
    return s3_client.generate_presigned_url(
        ClientMethod="put_object",
        Params={"Bucket": s3_bucket_name,
                "Key": object_name, "ContentType": content_type},
        ExpiresIn=expiration,
        HttpMethod="PUT",
    )


def __list_execution(event, group_name):
    max_items = __get_query_parameter(event, "MaxItems", DEFAULT_MAX_ITEMS)
    page_size = __get_query_parameter(event, "PageSize", DEFAULT_SIZE)
    starting_token = __get_query_parameter(event, "StartingToken")
    config = {
        "MaxItems": int(max_items),
        "PageSize": int(page_size),
        "StartingToken": starting_token,
    }
    response = dynamodb_client.query(
        TableName=intention_table_name,
        KeyConditionExpression='groupName = :groupName',
        ExpressionAttributeValues={
            ':groupName': {'S': group_name}
        }
    )
    output = {}
    page_json = []
    items = response['Items']
    while 'LastEvaluatedKey' in response:
        response = dynamodb_client.query(
            TableName=intention_table_name,
            KeyConditionExpression='groupName = :pk_val',
            ExpressionAttributeValues={
                ':pk_val': {'S': group_name}
            },
            ExclusiveStartKey=response['LastEvaluatedKey']
        )
        items.extend(response['Items'])

    for item in items:
        item_json = {}
        for key in list(item.keys()):
            value = item.get(key, {"S": ""}).get("S", "-")
            if key == "File":
                item_json["fileName"] = value.split("/").pop()
            elif key == "modelId":
                item_json["model"] = value
            elif key == "LastModifiedTime":
                item_json["createTime"] = value
            elif key == "LastModifiedBy":
                item_json["createBy"] = value
            elif key == "intentionId":
                item_json["executionId"] = value
            else:
                item_json[key] = value
            item_json["executionStatus"] = "COMPLETED"
        page_json.append(item_json)
        output["Items"] = page_json
    output["Config"] = config
    output["Count"] = len(items)
    return output


def __create_execution(event, context, email, group_name):
    input_body = json.loads(event["body"])
    execution_detail = {}
    execution_detail["tableItemId"] = context.aws_request_id
    execution_detail["chatbotId"] = input_body.get("chatbotId")
    execution_detail["groupName"] = group_name
    execution_detail["index"] = input_body.get("index")
    execution_detail["model"] = input_body.get("model")
    execution_detail["fileName"] = input_body.get("s3Prefix").split("/").pop()
    bucket = input_body.get("s3Bucket")
    prefix = input_body.get("s3Prefix")
    s3_response = __get_s3_object_with_retry(bucket, prefix)
    file_content = s3_response['Body'].read()
    excel_file = BytesIO(file_content)
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    qaList = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        question, intention, kwargs = row[0], row[1], row[2] if len(
            row) > 2 else None
        if not question:
            continue
        # for i, element in enumerate(qaList):
        #     if element.get("question") == question:
        #         qaList[i] = {
        #             "question": question,
        #             "intention": intention,
        #             "kwargs": kwargs
        #         }
        #         return qaList
        qaList.append({
            "question": question,
            "intention": intention,
            "kwargs": kwargs
        })
    # write to ddb(meta data)
    intention_table.put_item(
        Item={
            "groupName": group_name,
            "intentionId": context.aws_request_id,
            "chatbotId": input_body.get("chatbotId"),
            "model": execution_detail["model"],
            "index": execution_detail["index"],
            "tag": execution_detail["index"],
            "File": f'{bucket}{input_body.get("s3Prefix")}',
            "LastModifiedBy": email,
            "LastModifiedTime": re.findall(r'\[(.*?)\]', input_body.get("s3Prefix"))[0],
            "details": json.dumps(qaList)
        }
    )
    # write to aos(vectorData)
    __save_2_aos(input_body.get("model"),
                 execution_detail["index"], qaList, bucket, prefix)

    return {
        "execution_id": execution_detail["tableItemId"],
        "input_payload": execution_detail,
        "result": "success"
    }


def convert_qa_list(qa_list: list, bucket: str, prefix: str) -> List[Document]:
    doc_list = []
    for qa in qa_list:
        metadata_template = {
            "content_type": "paragraph",
            "heading_hierarchy": {},
            "figure_list": [],
            "chunk_id": "$$",
            "file_path": "",
            "keywords": [],
            "summary": "",
            "type": "Intent"
        }
        page_content = qa["question"]
        metadata = metadata_template
        metadata["jsonlAnswer"] = {}
        metadata["jsonlAnswer"]["answer"] = qa["intention"]
        metadata["jsonlAnswer"]["question"] = qa["question"]
        metadata["jsonlAnswer"]["intent"] = qa["intention"]
        metadata["kwargs"] = qa["kwargs"]
        metadata["file_path"] = f"s3://{bucket}/{prefix}"
        # Assemble the document
        doc = Document(page_content=page_content, metadata=metadata)
        doc_list.append(doc)

    return doc_list


def __save_2_aos(modelId: str, index: str, qaListParam: list, bucket: str, prefix: str):
    qaList = __deduplicate_by_key(qaListParam, "question")
    if kb_enabled and embedding_model_endpoint.startswith("bce-embedding"):
        embedding_info = get_embedding_info(embedding_model_endpoint)
        embedding_function = sm_utils.getCustomEmbeddings(
            embedding_model_endpoint, region, embedding_info.get("ModelType")
        )
        docsearch = OpenSearchVectorSearch(
            index_name=index,
            embedding_function=embedding_function,
            opensearch_url="https://{}".format(aosEndpoint),
            http_auth=awsauth,
            use_ssl=True,
            verify_certs=True,
            connection_class=RequestsHttpConnection,
        )

        worker = OpenSearchIngestionWorker(docsearch, embedding_model_endpoint)
        intention_list = convert_qa_list(qaList, bucket, prefix)
        worker.aos_ingestion(intention_list, index)
    else:
        index_exists = aos_client.indices.exists(index=index)
        if not index_exists:
            __create_index(index, modelId)
        __refresh_index(index, modelId, qaList)


def __create_index(index: str, modelId: str):
    body = {
        "settings": {
            "index": {
                "number_of_shards": 1,
                "number_of_replicas": 0,
                "knn": True,
                "knn.algo_param.ef_search": 32
            }
        },
        "mappings": {
            "properties": {
                "text": {
                    "type": "text",
                    "fields": {
                        "keyword": {
                            "type": "keyword"
                        }
                    }
                },
                "sentence_vector": {
                    "type": "knn_vector",
                    "dimension": ModelDimensionMap[modelId],
                    "method": {
                        "engine": "nmslib",
                        "space_type": "l2",
                        "name": "hnsw",
                        "parameters": {
                            "ef_construction": 512,
                            "m": 16
                        }
                    }
                }
            }
        }
    }
    try:
        aos_client.indices.create(index=index, body=body)
        logger.info("Index %s created successfully.", index)
    except RequestError as error:
        logger.info(error.error)


def __refresh_index(index: str, modelId: str, qaList):
    success, failed = helpers.bulk(aos_client,  __append_embeddings(
        index, modelId, qaList), chunk_size=BULK_SIZE)
    aos_client.indices.refresh(index=index)
    logger.info("Successfully added: %d ", success)
    logger.info("Failed: %d ", len(failed))


def __append_embeddings(index, modelId, qaList: list):
    actions = []
    documents = []
    for item in qaList:
        question = item["question"]
        embedding_func = BedrockEmbeddings(
            client=bedrock_client,
            model_id=modelId,
            normalize=True
        )

        embeddings_vectors = embedding_func.embed_documents(
            [question]
        )
        documents.append(
            {
                "text": question,
                "metadata": {
                    "answer": item["intention"],
                    "source": "portal",
                    **({"kwargs": item["kwargs"]} if item.get("kwargs") else {}),
                    "type": "Intent"
                },
                "sentence_vector": embeddings_vectors[0]
            }
        )

    for document in documents:
        index_list = index.split(",")
        for index_item in index_list:
            doc_id = hashlib.md5(str(document["text"]).encode('utf-8')).hexdigest()
            action = {
                "_op_type": "index",
                "_index": index_item,
                "_id": doc_id,
                "_source": document
            }
            actions.append(action)
    return actions
            # yield {"_op_type": "index", "_index": index_item, "_source": document, "_id": hashlib.md5(str(document).encode('utf-8')).hexdigest()}


def __get_execution(event, group_name):
    executionId = event.get("path", "").split("/").pop()
    index_response = intention_table.get_item(
        Key={
            "groupName": group_name,
            "intentionId": executionId,
        },
    )
    item = index_response['Item']
    res = {}
    Items = []
    # for item in items:
    item_json = {}
    for key in list(item.keys()):
        value = item.get(key)
        if key == "File":
            split_index = value.rfind('/')
            if split_index != -1:
                item_json["s3Path"] = value[:split_index]
                item_json["s3Prefix"] = value[split_index + 1:]
            else:
                item_json["s3Path"] = value
                item_json["s3Prefix"] = "-"
        elif key == "LastModifiedTime":
            item_json["createTime"] = value
        elif key == "details":
            item_json["QAList"] = json.loads(value)
        else:
            continue
        item_json["status"] = "COMPLETED"
    Items.append(item_json)
    res["Items"] = Items
    res["Count"] = len(Items)
    return res


def __get_s3_object_with_retry(bucket: str, key: str, max_retries: int = 5, delay: int = 1):
    attempt = 0
    while attempt < max_retries:
        try:
            return s3_client.get_object(Bucket=bucket, Key=key)
        except Exception:
            logger.info("Attempt %d failed", attempt + 1)
            attempt += 1
            if attempt >= max_retries:
                logger.info("Time out, retry...")
                raise
            time.sleep(delay)

def __download_template():
    url = s3_client.generate_presigned_url(
        ClientMethod="get_object",
        Params={'Bucket': s3_bucket_name,
                'Key': "templates/intention_corpus.xlsx"},
        ExpiresIn=60
    )
    return url


def __index_used_scan(event, group_name):
    input_body = json.loads(event["body"])
    index_response = index_table.get_item(
        Key={
            "groupName": group_name,
            "indexId": input_body.get("index"),
        },
    )
    pre_model = index_response.get("Item")
    model_name = ''
    if pre_model:
        model_response = model_table.get_item(
            Key={
                "groupName": group_name,
                "modelId": pre_model.get("modelIds", {}).get("embedding"),
            }
        )
        model_name = model_response.get("Item", {}).get(
            "parameter", {}).get("ModelName", "")
        #  model_name = model_response.get("ModelName", {}).get("S","-")
    if not pre_model or model_name == input_body.get("model"):
        return {
            "statusCode": 200,
            "headers": resp_header,
            "body": json.dumps({
                "result": "valid"
            })
        }
    else:
        return {
            "statusCode": 200,
            "headers": resp_header,
            "body": json.dumps({
                "result": "invalid"
            }
            )}


def __deduplicate_by_key(lst, key):
    seen = {}
    for element in lst:
        seen[element[key]] = element
    return list(seen.values())

def __get_query_parameter(event, parameter_name, default_value=None):
    if (
        event.get("queryStringParameters")
        and parameter_name in event["queryStringParameters"]
    ):
        return event["queryStringParameters"][parameter_name]
    return default_value
